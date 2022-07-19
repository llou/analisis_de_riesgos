#! /usr/bin/env python

from yaml import load, CLoader
from openpyxl import Workbook


POTENCIAL_AMENAZA = [
        [ 0, 0, 0, 0 ],
        [ 0, 1, 1, 2],
        [ 0, 1, 2, 3]]


def potencial_amenaza(vulnerabilidad, amenaza):
    return POTENCIAL_AMENAZA[vulnerabilidad][amenaza]


class Amenaza:
    def __init__(self, amenaza):
        self.nombre = amenaza['nombre']
        self.codigo = amenaza['codigo']
        self.nivel = amenaza['nivel']


class Amenazas:
    def __init__(self):
        amenazas = load(open('amenazas.yml'), Loader=CLoader)
        self.amenazas = [Amenaza(x) for x in amenazas]

    def __getitem__(self, codigo):
        for item in self.amenazas:
            if item.codigo == codigo:
                return item
        print(codigo)
        raise KeyError

amenazas = Amenazas()

def max_vectors(*vectors):
    result = []
    for i in range(0, 12):
        max = 0
        for vector in vectors:
            if vector[i] > max:
                max = vector[i]
        result.append(max)
    return result       

class Impactos:
    etiquetas_disponibilidad = ['Interactividad','1hora','1/2 día', '1 día',
                                '1 semana', '2 semanas+']
    etiquetas_confidencialidad = ['Interna', 'Externa']
    etiquetas_integridad = ['Perdida', 'Error', 'Manipulación', 'Autenticidad']
    etiquetas = etiquetas_disponibilidad + etiquetas_confidencialidad + etiquetas_integridad 
    
    @classmethod
    def max(cls, *impactos):
        vectors = [riesgo.vector for riesgo in impactos]
        max_vector = max_vectors(*vectors)
        return cls(max_vector)


    def __init__(self, vector):
        self.vector = vector

    def resumen(self):
        return [max(self.vector[0:7]), 
                max(self.vector[7:9]), 
                max(self.vector[9:])]

    def a_celdas(self, ws, row, col, resumen=True):
        for i, v in enumerate(self.vector):
            ws.cell(row=row, column=col + i, value=v)
        if resumen:
            resumen = self.resumen()
            ws.cell(row=row, column=col + 12, value=resumen[0])
            ws.cell(row=row, column=col + 13, value=resumen[1])
            ws.cell(row=row, column=col + 14, value=resumen[2])

    @property
    def disponibilidad(self):
        return dict(self.etiquetas_disponibilidad, self.vector[0,6])

    @property
    def confidencialidad(self):
        return dict(self.etiquetas_confidencialidad, self.vector[6,8])

    @property
    def Integridad(self):
        return dict(self.etiquetas_integridad, self.vector[8,12])

    def __int__(self):
        return max(self.vector)

IMPACTO_NULO = Impactos([0,0,0,0,0,0,0,0,0,0,0,0])


class Activo:
    def __init__(self, activo, activos):
        self.activo = activo
        self.activos = activos

    @property
    def nombre(self):
        return self.activo['nombre']

    @property
    def tipo(self):
        return self.activo['tipo']

    @property
    def codigo(self):
        return self.activo['codigo']

    @property
    def dependencias(self):
        if 'dependencias' in self.activo:
            return self.activo['dependencias']
        return []

    def itera_padres(self):
        for activo in self.activos:
            if self.codigo in activo.dependencias:
                yield activo
                for hijo in activo.itera_padres():
                    yield hijo


class ActivoDeNegocio(Activo):
    def __init__(self, activo, activos):
        super().__init__(activo, activos)
        if 'impactos' in activo:
            self.impactos = Impactos(activo['impactos'])
        else:
            self.impactos = IMPACTO_NULO


class ActivoTecnologico(Activo):
    def __init__(self, activo, activos):
        self.activos = activos
        self.activo = activo

    @property
    def codigo(self):
        return self.activo['codigo']

    def activos_de_negocio(self):
        ans = []
        for padre in self.itera_padres():
            if isinstance(padre, ActivoDeNegocio):
                ans.append(padre)
        return ans

    @property
    def impacto(self):
        impactos = [a.impactos for a in self.activos_de_negocio()]
        return Impactos.max(*impactos)

    @property
    def potencial_amenaza(self):
        valor_potencial_amenaza = 0
        for vul in self.activo['vulnerabilidades']:
            codigo = vul['amenaza']
            vulnerabilidad = vul['vulnerabilidad']
            amenaza = amenazas[codigo]
            nivel_amenaza = amenaza.nivel
            p_a = potencial_amenaza(vulnerabilidad, nivel_amenaza)
            if p_a > valor_potencial_amenaza:
                valor_potencial_amenaza = p_a
        return valor_potencial_amenaza

    @property
    def riesgo(self):
        return self.potencial_amenaza * int(self.impacto)

    def itera_vulnerabilidades(self):
        for vul in self.activo['vulnerabilidades']:
            vulnerabilidad = vul['vulnerabilidad']
            codigo = vul['amenaza']
            amenaza = amenazas[codigo]
            nivel_amenaza = amenaza.nivel
            p_a = potencial_amenaza(vulnerabilidad, nivel_amenaza)
            yield dict(nombre_amenaza=amenaza.nombre, 
                       vulnerabilidad=vulnerabilidad,
                       nivel_amenaza=nivel_amenaza,
                       potencial_amenaza=p_a)



class Activos:
    def __init__(self):
        activos = load(open('activos.yml'), Loader=CLoader)
        self.capa_negocio = []
        self.capa_aplicacion = []
        self.capa_tecnologica = []
        self.indice = {}

        for activo in activos:
            if activo['tipo'] in ['Proceso', 'Servicio', 'Rol', 'Información']:
                negocio = ActivoDeNegocio(activo, self)
                self.capa_negocio.append(negocio)
                self.indice[negocio.codigo] = negocio

            elif activo['tipo'] in ['Aplicación', 'Base de datos']:
                activo = ActivoTecnologico(activo, self)
                self.capa_aplicacion.append(activo)
                self.indice[activo.codigo] = activo
            else:
                activo = ActivoTecnologico(activo, self)
                self.capa_tecnologica.append(activo)
                self.indice[activo.codigo] = activo

        self.tecnologicos = self.capa_aplicacion + self.capa_tecnologica
        self.activos = self.capa_negocio + self.tecnologicos

    def itera_amenazas(self):
        for a_t in self.tecnologicos:
            for vul in a_t.itera_vulnerabilidades():
                vul['nombre'] = a_t.nombre
                vul['tipo'] = a_t.tipo
                yield vul

    def __iter__(self):
        return self.activos.__iter__()

activos = Activos()

wb = Workbook()

dest_filename = "riesgos.xlsx"

ws1 = wb.active
ws1.title = "Activos"

S = 2

for i, activo in enumerate(activos.itera_amenazas()):
    ws1.cell(row=S+i+1, column=1, value=activo['nombre'])
    ws1.cell(row=S+i+1, column=2, value=activo['tipo'])
    ws1.cell(row=S+i+1, column=3, value=activo['nombre_amenaza'])
    ws1.cell(row=S+i+1, column=4, value=activo['nivel_amenaza'])
    ws1.cell(row=S+i+1, column=5, value=activo['vulnerabilidad'])
    ws1.cell(row=S+i+1, column=6, value=activo['potencial_amenaza'])

ws2 = wb.create_sheet("Propagación de impactos")

S = 3

ws2['C2'] = "Disponibilidad"
ws2['I2'] = "Confidencialidad"
ws2['K2'] = "Integridad"
for i, etiqueta in enumerate(Impactos.etiquetas):
    ws2.cell(row=S, column=i + 3, value=etiqueta)

ws2['C2'] = "Disponibilidad"
ws2['I2'] = "Confidencialidad"
ws2['K2'] = "Integridad"
ws2.cell(row=S, column=15, value="Disponibilidad")
ws2.cell(row=S, column=16, value="Confidencialidad")
ws2.cell(row=S, column=17, value="Integridad")
ws2.cell(row=S, column=18, value="Impacto")

for i, activo in enumerate(activos.tecnologicos):
    ws2.cell(row=S + i+1, column=1, value=activo.nombre)
    ws2.cell(row=S+i+1, column=2, value=activo.tipo)
    activo.impacto.a_celdas(ws2, S+i+1, 3, resumen=True)
    ws2.cell(row=S+i+1, column=18, value=int(activo.impacto))

ws3 = wb.create_sheet("Cálculo de riesgos")

S = 3

ws3.cell(row=S, column=3, value="Impacto")
ws3.cell(row=S, column=4, value="Potencial Amenaza")
ws3.cell(row=S, column=5, value="Riesgo")

for i, activo in enumerate(activos.tecnologicos):
    ws3.cell(row=S + i+1, column=1, value=activo.nombre)
    ws3.cell(row=S+i+1, column=2, value=activo.tipo)
    ws3.cell(row=S+i+1, column=3, value=int(activo.impacto))
    ws3.cell(row=S+i+1, column=4, value=activo.potencial_amenaza)
    ws3.cell(row=S+i+1, column=5, value=activo.riesgo)

wb.save(filename = dest_filename)

# Activos
